"""
Genera un Excel formateado profesionalmente a partir del JSON de análisis.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta ────────────────────────────────────────────────────────────────────
TEAL      = '0F6E56'
TEAL2     = '1D9E75'
BLUE_ESC1 = '1E3A8A'
PETROL    = '0F766E'
GREEN_ESC = '166534'
WHITE     = 'FFFFFF'
GRAY_ROW  = 'F5F4F0'
GRAY_HDR  = 'E8E7E2'

def _font(bold=False, color='000000', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)

def _fill(color):
    return PatternFill('solid', fgColor=color)

def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _title(ws, row, cols, bg=TEAL, fg=WHITE, size=12):
    ws.cell(row=row, column=1).font = _font(bold=True, color=fg, size=size)
    ws.cell(row=row, column=1).fill = _fill(bg)
    ws.cell(row=row, column=1).alignment = _align('left')
    for c in range(2, cols+1):
        ws.cell(row=row, column=c).fill = _fill(bg)

def _header(ws, row, cols, bg, fg=WHITE):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = _font(bold=True, color=fg, size=9)
        cell.fill = _fill(bg)
        cell.alignment = _align('center', wrap=True)
        cell.border = _border()

def _semcolor(val):
    s = str(val) if val is not None else ''
    if '🔴' in s: return ('FECACA', '7F1D1D')
    if '🟡' in s: return ('FEF08A', '854D0E')
    if '🟢' in s: return ('BBF7D0', '166534')
    return None, None

def _data_row(ws, row, cols, alt=False, bold_col1=False):
    bg = GRAY_ROW if alt else WHITE
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        bg_sem, fg_sem = _semcolor(cell.value)
        if bg_sem:
            cell.fill = _fill(bg_sem)
            cell.font = _font(bold=True, size=9, color=fg_sem)
        else:
            cell.fill = _fill(bg)
            cell.font = _font(bold=(bold_col1 and c==1), size=9)
            cell.alignment = _align('left') if c==1 else _align('center')
        cell.border = _border()

def _freeze(ws, row=5, col=2):
    ws.freeze_panes = ws.cell(row=row, column=col)
    ws.sheet_view.showGridLines = False

def d(v):
    return v if v is not None else '—'

def semR(v):
    if not isinstance(v, (int, float)): return '—'
    s = f'{v:.2f}'
    if v >= 4:   return f'🔴 {s}'
    if v >= 2.5: return f'🟡 {s}'
    return f'🟢 {s}'

def semP(v):
    if not isinstance(v, (int, float)): return '—'
    s = f'{v*100:.1f}%'
    if v >= 0.60: return f'🔴 {s}'
    if v >= 0.35: return f'🟡 {s}'
    return f'🟢 {s}'

def semPS(v):
    if not isinstance(v, (int, float)): return '—'
    s = f'{v*100:.1f}%'
    if v >= 0.75: return f'🔴 {s}'
    if v >= 0.25: return f'🟡 {s}'
    return f'🟢 {s}'


def generar_excel_formateado(data: dict, periodo: str, centro: str) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    e1 = data.get('escenario_1', {})
    e2 = data.get('escenario_2', {})
    e3 = data.get('escenario_3', {})

    # ── Hoja 1: Resumen Ejecutivo ─────────────────────────────────────────────
    ws = wb.create_sheet('Resumen Ejecutivo')
    rows = [
        [f'ANÁLISIS DE PRESCRIPCIONES POR CONSULTA — {periodo}'],
        [f'Centro: {centro} | 28 categorías | PAMI SDLC + PAMI OFTAL'],
        ['RESUMEN COMPARATIVO POR ESCENARIO'],
        ['Escenario','Consultas','Con Presc','Sin Presc','% Con Presc','Total Presc','Ratio P/C'],
        ['ESC 1 — COMPLETO', d(e1.get('total_consultas')), d(e1.get('con_prescripcion')), d(e1.get('sin_prescripcion')), semP((e1.get('pct_con_presc') or 0)/100), d(e1.get('total_prescripciones')), semR(e1.get('ratio'))],
        ['ESC 2 — SIN PAMI', d(e2.get('total_consultas')), d(e2.get('con_prescripcion')), d(e2.get('sin_prescripcion')), semP((e2.get('pct_con_presc') or 0)/100), d(e2.get('total_prescripciones')), semR(e2.get('ratio'))],
        ['ESC 3 — SIN PAMI NI LAB', d(e3.get('total_consultas')), d(e3.get('con_prescripcion')), d(e3.get('sin_prescripcion')), semP((e3.get('pct_con_presc') or 0)/100), d(e3.get('total_prescripciones')), semR(e3.get('ratio'))],
        [],
        ['NOTAS METODOLÓGICAS'],
        ['• Denominador: todas las consultas (n_presc=0 si sin prescripción)'],
        ['• Join: Paciente + Profesional + Fecha'],
        ['• Ratio = Total Prescripciones / Total Consultas'],
        ['• PAMI: incluye PAMI SDLC y PAMI OFTAL'],
        ['• Excluye vacuna antigripal y prestaciones de vacunación'],
        ['• Semáforo Ratio: 🟢<2.5 | 🟡2.5-4 | 🔴≥4'],
        ['• Semáforo %ConPresc: 🟢<35% | 🟡35-60% | 🔴≥60%'],
        [],
        [data.get('notas', '')],
    ]
    for row in rows:
        ws.append(row)

    _title(ws, 1, 7, TEAL, WHITE, 13)
    _title(ws, 2, 7, TEAL2, WHITE, 9)
    _title(ws, 3, 7, BLUE_ESC1, WHITE, 10)
    _header(ws, 4, 7, '374151')

    esc_colors = {'ESC 1': BLUE_ESC1, 'ESC 2': PETROL, 'ESC 3': GREEN_ESC}
    for r in range(5, ws.max_row+1):
        v = ws.cell(row=r, column=1).value
        sv = str(v) if v else ''
        if any(k in sv for k in esc_colors):
            bg = next(v2 for k, v2 in esc_colors.items() if k in sv)
            ws.cell(row=r, column=1).font = _font(bold=True, color=WHITE, size=9)
            ws.cell(row=r, column=1).fill = _fill(bg)
            ws.cell(row=r, column=1).alignment = _align('left')
            for c in range(2, 8):
                cell = ws.cell(row=r, column=c)
                bg_s, fg_s = _semcolor(cell.value)
                if bg_s:
                    cell.fill = _fill(bg_s)
                    cell.font = _font(bold=True, size=9, color=fg_s)
                else:
                    cell.fill = _fill('F0FDF4' if 'ESC 3' in sv else 'EFF6FF' if 'ESC 1' in sv else 'F0FDFA')
                    cell.font = _font(bold=True, size=9)
                cell.alignment = _align('center')
                cell.border = _border()
        elif 'NOTAS' in sv.upper():
            _title(ws, r, 7, GRAY_HDR, '000000', 9)
        elif sv.startswith('•'):
            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.font = _font(size=9, italic=True, color='444444')
                cell.fill = _fill('FAFAF8')
                cell.alignment = _align('left')

    ws.column_dimensions['A'].width = 30
    for col in 'BCDEFG':
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[1].height = 22
    _freeze(ws, 5, 2)

    # ── Hoja 2: Por Servicio ──────────────────────────────────────────────────
    ws = wb.create_sheet('Por Servicio')
    hdr_svc = ['Servicio','Consultas','Con Presc','Sin Presc','% Con Presc','Total Presc','Ratio P/C',
                'Consultas','Con Presc','Sin Presc','% Con Presc','Total Presc','Ratio P/C',
                'Consultas','Con Presc','Sin Presc','% Con Presc','Total Presc','Ratio P/C']
    ws.append([f'PRESCRIPCIONES POR CONSULTA — POR SERVICIO — {periodo}'])
    ws.append(['3 escenarios lado a lado'])
    ws.append(['','[ESC 1: COMPLETO]','','','','','','[ESC 2: SIN PAMI]','','','','','','[ESC 3: SIN PAMI NI LAB]','','','','',''])
    ws.append(hdr_svc)

    svcs = sorted(data.get('por_servicio', []), key=lambda x: -(x.get('esc1_consultas') or 0))
    for s in svcs:
        ws.append([
            s.get('servicio',''),
            d(s.get('esc1_consultas')), d(s.get('esc1_con_presc')), d(s.get('esc1_sin_presc')), semP((s.get('esc1_pct') or 0)/100), d(s.get('esc1_total_presc')), semR(s.get('esc1_ratio')),
            d(s.get('esc2_consultas')), d(s.get('esc2_con_presc')), d(s.get('esc2_sin_presc')), semP((s.get('esc2_pct') or 0)/100), d(s.get('esc2_total_presc')), semR(s.get('esc2_ratio')),
            d(s.get('esc3_consultas')), d(s.get('esc3_con_presc')), d(s.get('esc3_sin_presc')), semP((s.get('esc3_pct') or 0)/100), d(s.get('esc3_total_presc')), semR(s.get('esc3_ratio')),
        ])

    CM = 19
    _title(ws, 1, CM, TEAL, WHITE, 12)
    _title(ws, 2, CM, TEAL2, WHITE, 9)
    for c in range(1, CM+1):
        v = ws.cell(row=3, column=c).value
        bg = BLUE_ESC1 if v and 'ESC 1' in str(v) else PETROL if v and 'ESC 2' in str(v) else GREEN_ESC if v and 'ESC 3' in str(v) else '374151'
        cell = ws.cell(row=3, column=c)
        cell.fill = _fill(bg); cell.font = _font(bold=True, color=WHITE, size=9); cell.alignment = _align('center')
    _header(ws, 4, CM, '374151')

    alt = False
    for r in range(5, ws.max_row+1):
        v1 = ws.cell(row=r, column=1).value
        if v1: alt = not alt
        _data_row(ws, r, CM, alt, bold_col1=bool(v1))

    ws.column_dimensions['A'].width = 28
    for c in range(2, CM+1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    _freeze(ws, 5, 2)

    # ── Hoja 3: Por Profesional ───────────────────────────────────────────────
    ws = wb.create_sheet('Por Profesional')
    hdr_prof = ['Profesional','Servicio',
                'Consultas','Con Presc','Sin Presc','% Con Presc','Total Presc','Ratio P/C',
                'Consultas','Con Presc','Sin Presc','% Con Presc','Total Presc','Ratio P/C',
                'Consultas','Con Presc','Sin Presc','% Con Presc','Total Presc','Ratio P/C']
    ws.append([f'PRESCRIPCIONES POR CONSULTA — POR PROFESIONAL — {periodo}'])
    ws.append(['','','[ESC 1: COMPLETO]','','','','','','[ESC 2: SIN PAMI]','','','','','','[ESC 3: SIN PAMI NI LAB]','','','','',''])
    ws.append(hdr_prof)

    profs = sorted(data.get('por_profesional', []), key=lambda x: -(x.get('esc1_consultas') or 0))
    for p in profs:
        ws.append([
            p.get('profesional',''), p.get('servicio',''),
            d(p.get('esc1_consultas')), d(p.get('esc1_con_presc')), d(p.get('esc1_sin_presc')), semP((p.get('esc1_pct') or 0)/100), d(p.get('esc1_total_presc')), semR(p.get('esc1_ratio')),
            d(p.get('esc2_consultas')), d(p.get('esc2_con_presc')), d(p.get('esc2_sin_presc')), semP((p.get('esc2_pct') or 0)/100), d(p.get('esc2_total_presc')), semR(p.get('esc2_ratio')),
            d(p.get('esc3_consultas')), d(p.get('esc3_con_presc')), d(p.get('esc3_sin_presc')), semP((p.get('esc3_pct') or 0)/100), d(p.get('esc3_total_presc')), semR(p.get('esc3_ratio')),
        ])

    CM = 20
    _title(ws, 1, CM, TEAL, WHITE, 12)
    for c in range(1, CM+1):
        v = ws.cell(row=2, column=c).value
        bg = BLUE_ESC1 if v and 'ESC 1' in str(v) else PETROL if v and 'ESC 2' in str(v) else GREEN_ESC if v and 'ESC 3' in str(v) else '374151'
        cell = ws.cell(row=2, column=c)
        cell.fill = _fill(bg); cell.font = _font(bold=True, color=WHITE, size=9); cell.alignment = _align('center')
    _header(ws, 3, CM, '374151')

    alt = False
    for r in range(4, ws.max_row+1):
        v1 = ws.cell(row=r, column=1).value
        if v1: alt = not alt
        _data_row(ws, r, CM, alt, bold_col1=bool(v1))
        ws.cell(row=r, column=2).alignment = _align('left')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    for c in range(3, CM+1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    _freeze(ws, 4, 3)

    # ── Hoja 4: Tipos x Servicio (Esc.3) ─────────────────────────────────────
    ws = wb.create_sheet('Tipos x Servicio (Esc.3)')
    ws.append([f'TIPOS DE ESTUDIO/PRÁCTICA POR SERVICIO — ESC 3: SIN PAMI NI LAB'])
    ws.append([f'{periodo} | {centro}'])
    ws.append(['Servicio','Tipo de Estudio / Práctica','Consultas Episodio','Prescripciones','Ratio P/C'])

    tipos_svc = sorted(data.get('tipos_por_servicio_esc3', []),
                       key=lambda x: (x.get('servicio',''), -(x.get('prescripciones') or 0)))
    last_svc = None
    for t in tipos_svc:
        svc = t.get('servicio','')
        ws.append([svc if svc != last_svc else '', t.get('tipo',''), d(t.get('consultas_episodio')), d(t.get('prescripciones')), semR(t.get('ratio'))])
        last_svc = svc

    _title(ws, 1, 5, GREEN_ESC, WHITE, 12)
    _title(ws, 2, 5, '166534', WHITE, 9)
    _header(ws, 3, 5, '374151')
    alt = False; last = None
    for r in range(4, ws.max_row+1):
        v1 = ws.cell(row=r, column=1).value
        if v1 and v1 != last: last = v1; alt = not alt
        _data_row(ws, r, 5, alt, bold_col1=bool(v1))
        ws.cell(row=r, column=2).alignment = _align('left')

    ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 18; ws.column_dimensions['D'].width = 16; ws.column_dimensions['E'].width = 14
    _freeze(ws, 4, 2)

    # ── Hoja 5: Tipos x Profesional (Esc.3) ──────────────────────────────────
    ws = wb.create_sheet('Tipos x Profesional (Esc.3)')
    ws.append([f'TIPOS DE ESTUDIO/PRÁCTICA POR PROFESIONAL — ESC 3'])
    ws.append([f'{periodo} | {centro} | 🟢%SinPresc<25% 🟡25-50% 🔴≥75%'])
    ws.append(['Profesional','Servicio','Tipo / Resumen','Consultas Episodio','Prescripciones','Ratio P/C','% Sin Presc'])

    for t in data.get('tipos_por_profesional_esc3', []):
        sin = semPS((t.get('sin_presc_pct') or 0)/100) if t.get('es_resumen') else ''
        ws.append([t.get('profesional',''), t.get('servicio',''), t.get('tipo',''),
                   d(t.get('consultas_episodio')), d(t.get('prescripciones')), semR(t.get('ratio')), sin])

    _title(ws, 1, 7, GREEN_ESC, WHITE, 12)
    _title(ws, 2, 7, '166534', WHITE, 9)
    _header(ws, 3, 7, '374151')
    alt = False; last = None
    for r in range(4, ws.max_row+1):
        v1 = ws.cell(row=r, column=1).value
        if v1 and v1 != last: last = v1; alt = not alt
        _data_row(ws, r, 7, alt, bold_col1=bool(v1))
        ws.cell(row=r, column=2).alignment = _align('left')
        ws.cell(row=r, column=3).alignment = _align('left')

    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 34
    for c in range(4, 8):
        ws.column_dimensions[get_column_letter(c)].width = 14
    _freeze(ws, 4, 3)

    # ── Hoja 6: Duración vs Prescripción ─────────────────────────────────────
    ws = wb.create_sheet('Duracion vs Prescripcion')
    ws.append([f'DURACIÓN DE CONSULTA vs. PRESCRIPCIÓN'])
    ws.append([f'{periodo} | {centro} | Excluidos: <5 min y >90 min'])
    ws.append(['Rango Duración','N° Consultas','Con Presc','Sin Presc','% Con Presc','Ratio P/C','Duración Prom'])

    for dv in data.get('duracion_vs_prescripcion', []):
        dp = dv.get('duracion_prom')
        ws.append([dv.get('rango',''), d(dv.get('consultas')), d(dv.get('con_presc')), d(dv.get('sin_presc')),
                   semP((dv.get('pct_con_presc') or 0)/100), semR(dv.get('ratio')),
                   f"{dp:.0f} min" if isinstance(dp, (int,float)) else '—'])

    _title(ws, 1, 7, TEAL, WHITE, 12)
    _title(ws, 2, 7, TEAL2, WHITE, 9)
    _header(ws, 3, 7, '374151')
    for r in range(4, ws.max_row+1):
        _data_row(ws, r, 7, (r-4)%2==0)

    ws.column_dimensions['A'].width = 16
    for col in 'BCDEFG':
        ws.column_dimensions[col].width = 14
    _freeze(ws, 4, 2)

    # ── Hoja 7: Desvíos y Oportunidades ──────────────────────────────────────
    ws = wb.create_sheet('Desvios y Oportunidades')
    ws.append([f'DESVÍOS Y OPORTUNIDADES DE MEJORA — {periodo}'])
    ws.append([f'Centro: {centro}'])
    ws.append([])
    ws.append(['TOP 12 PROFESIONALES POR RATIO — Escenario 3 (Sin PAMI ni Laboratorio)'])
    ws.append([])
    ws.append(['Profesional','Servicio','Ratio P/C','Consultas','Total Presc'])

    for dv in data.get('desvios_top12', []):
        ws.append([dv.get('profesional',''), dv.get('servicio',''), semR(dv.get('esc3_ratio')), d(dv.get('esc3_consultas')), d(dv.get('esc3_total_presc'))])

    sin_presc = data.get('servicios_sin_prescripcion_esc3', [])
    if sin_presc:
        ws.append([]); ws.append([]); ws.append(['SERVICIOS SIN PRESCRIPCIONES — Escenario 3']); ws.append([])
        for s in sin_presc:
            ws.append([s])

    if data.get('notas'):
        ws.append([]); ws.append(['HALLAZGOS Y NOTAS']); ws.append([data['notas']])

    _title(ws, 1, 5, TEAL, WHITE, 12)
    _title(ws, 2, 5, TEAL2, WHITE, 9)
    for r in range(3, ws.max_row+1):
        v1 = ws.cell(row=r, column=1).value
        sv = str(v1) if v1 else ''
        if 'TOP' in sv or 'SERVICIOS' in sv or 'HALLAZGOS' in sv:
            _title(ws, r, 5, BLUE_ESC1, WHITE, 10)
        elif sv == 'Profesional':
            _header(ws, r, 5, '374151')
        elif v1:
            _data_row(ws, r, 5, r%2==0, bold_col1=False)

    ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 12; ws.column_dimensions['D'].width = 12; ws.column_dimensions['E'].width = 14
    _freeze(ws, 7, 2)

    # ── Guardar en memoria ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
