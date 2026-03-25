"""
Genera el Excel de salida con los resultados del análisis.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

VERDE    = "FF1D9E75"
AMARILLO = "FFBA7517"
ROJO     = "FFE24B4A"
GRIS     = "FF888780"
TEAL_BG  = "FFE1F5EE"
PURP_BG  = "FFEEEDFE"
HEADER_BG = "FF0F6E56"
HEADER_FG = "FFFFFFFF"

def _estilo_header(ws, fila: int, cols: list[str]):
    for i, titulo in enumerate(cols, 1):
        c = ws.cell(row=fila, column=i, value=titulo)
        c.font      = Font(bold=True, color=HEADER_FG, size=10)
        c.fill      = PatternFill("solid", fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _borde_fino():
    lado = Side(style="thin", color="FFD3D1C7")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _color_ratio(ratio) -> str | None:
    if ratio is None:
        return None
    if ratio >= 4:
        return ROJO
    if ratio >= 2.5:
        return AMARILLO
    return None

def _autowidth(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            max(min_w, min(max_w, max_len + 3))

def _escribir_tabla(ws, datos: list[dict], fila_inicio: int,
                    cols_orden: list[str] | None = None) -> int:
    if not datos:
        ws.cell(row=fila_inicio, column=1, value="Sin datos para este escenario")
        return fila_inicio + 1

    cols = cols_orden or list(datos[0].keys())
    _estilo_header(ws, fila_inicio, cols)

    for r, row in enumerate(datos, fila_inicio + 1):
        for c_idx, col in enumerate(cols, 1):
            val = row.get(col)
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = _borde_fino()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == "ratio":
                color = _color_ratio(val)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(color="FFFFFFFF", bold=True)
            if col == "sin_prescripcion" and val is True:
                cell.fill = PatternFill("solid", fgColor=GRIS)

    return r + 2


def generar_excel(resultado: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _hoja_resumen(wb, resultado)
    for esc_key, titulo_hoja in [
        ("escenario_1", "Esc1 — Completo"),
        ("escenario_2", "Esc2 — Sin PAMI"),
        ("escenario_3", "Esc3 — Sin PAMI ni Lab"),
    ]:
        esc = resultado.get(esc_key, {})
        _hoja_escenario(wb, esc, titulo_hoja)

    _hoja_sin_prescripcion(wb, resultado)
    _hoja_desvios(wb, resultado)

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _hoja_resumen(wb: Workbook, resultado: dict):
    ws = wb.create_sheet("Resumen")
    meta = resultado.get("metadata", {})

    ws["A1"] = "Análisis de Conducta Prescriptiva"
    ws["A1"].font = Font(bold=True, size=14, color="FF0F6E56")
    ws.merge_cells("A1:G1")

    ws["A3"] = "Profesionales únicos:"
    ws["B3"] = meta.get("profesionales_unicos", "")
    ws["A4"] = "Total consultas (bruto):"
    ws["B4"] = meta.get("total_consultas_bruto", "")
    ws["A5"] = "Total prescripciones (bruto):"
    ws["B5"] = meta.get("total_prescripciones_bruto", "")

    ws["A7"] = "Centros incluidos:"
    ws["B7"] = ", ".join(meta.get("centros", []))

    fila = 10
    ws.cell(row=fila, column=1, value="Comparativo por escenario").font = \
        Font(bold=True, size=11)
    fila += 1

    cols = ["Escenario", "Consultas", "Prescripciones",
            "Con prescripción", "Sin prescripción", "Ratio global"]
    _estilo_header(ws, fila, cols)
    fila += 1

    for esc_key in ["escenario_1", "escenario_2", "escenario_3"]:
        esc = resultado.get(esc_key, {})
        vals = [
            esc.get("nombre", ""),
            esc.get("total_consultas", 0),
            esc.get("total_prescripciones", 0),
            esc.get("profesionales_con_prescripcion", 0),
            esc.get("profesionales_sin_prescripcion", 0),
            esc.get("ratio_global", 0),
        ]
        for c_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=fila, column=c_idx, value=v)
            cell.border = _borde_fino()
            cell.alignment = Alignment(horizontal="center")
        fila += 1

    _autowidth(ws)


def _hoja_escenario(wb: Workbook, esc: dict, titulo: str):
    ws = wb.create_sheet(titulo)

    ws["A1"] = f"Por profesional — {esc.get('nombre', titulo)}"
    ws["A1"].font = Font(bold=True, size=12, color="FF0F6E56")

    cols_prof = ["personal", "servicio", "centro",
                 "n_consultas", "n_prescripciones", "ratio", "sin_prescripcion"]
    fila = _escribir_tabla(ws, esc.get("por_profesional", []),
                           fila_inicio=3, cols_orden=cols_prof)

    ws.cell(row=fila, column=1, value="Por servicio").font = Font(bold=True)
    fila = _escribir_tabla(ws, esc.get("por_servicio", []),
                           fila_inicio=fila + 1)

    ws.cell(row=fila, column=1, value="Por centro").font = Font(bold=True)
    _escribir_tabla(ws, esc.get("por_centro", []), fila_inicio=fila + 1)

    _autowidth(ws)


def _hoja_sin_prescripcion(wb: Workbook, resultado: dict):
    ws = wb.create_sheet("Sin prescripción")
    ws["A1"] = "Profesionales que atendieron pero no prescribieron nada"
    ws["A1"].font = Font(bold=True, size=12, color="FF5F5E5A")

    fila = 3
    for esc_key, label in [
        ("escenario_1", "Escenario 1 — Completo"),
        ("escenario_3", "Escenario 3 — Sin PAMI ni Lab"),
    ]:
        ws.cell(row=fila, column=1, value=label).font = Font(bold=True)
        sin_presc = resultado.get(esc_key, {}).get("sin_prescripcion", [])
        cols = ["personal", "servicio", "centro", "n_consultas"]
        fila = _escribir_tabla(ws, sin_presc, fila_inicio=fila + 1,
                               cols_orden=cols)

    _autowidth(ws)


def _hoja_desvios(wb: Workbook, resultado: dict):
    ws = wb.create_sheet("Desvíos")
    ws["A1"] = "Top prescriptores por ratio — Escenario 3 (sin PAMI ni Lab)"
    ws["A1"].font = Font(bold=True, size=12, color="FFA32D2D")

    ws["A3"] = "Referencia de colores:"
    ws.cell(row=4, column=1, value="Normal (ratio < media + 1σ)").fill = \
        PatternFill("solid", fgColor="FFE1F5EE")
    ws.cell(row=5, column=1, value="Atención (ratio ≥ media + 1σ)").fill = \
        PatternFill("solid", fgColor=AMARILLO)
    ws.cell(row=6, column=1, value="Crítico (ratio ≥ media + 2σ)").fill = \
        PatternFill("solid", fgColor=ROJO)

    cols = ["personal", "servicio", "centro",
            "n_consultas", "n_prescripciones", "ratio", "z_score", "alerta"]
    _escribir_tabla(ws,
                    resultado.get("escenario_3", {}).get("top_desvios", []),
                    fila_inicio=8, cols_orden=cols)
    _autowidth(ws)
